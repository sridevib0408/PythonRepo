import win32com.client
import sys
import re
import pandas
from openpyxl import load_workbook,Workbook
import datetime
from os import path

class EmailReader(object):

    def __init__(self):
        self.outlook = win32com.client.Dispatch("Outlook.Application").GetNamespace("MAPI")
        try:
            root_folder = self.outlook.Folders.Item(1)
            # Get all mails in Alerts folder
            self.messages = root_folder.Folders['Alerts'].items
        except Exception as e:
            print("Unable to open email " + str(e))
            sys.exit(-1)

    #clean up
    def close(self):
        del self.outlook

    #Return all messages
    def getMessages(self):
        return self.messages

    def processEmailRequest(self,datefrm,dateto):
    

        #create file it doesnt exists
        wb = Workbook()
        fileName = 'C:\temp\Alerts_log.xlsx'
        if not path.exists(fileName):
            wb.save(fileName)
            del wb
	
        df = pandas.DataFrame()
        dateExcel = str(datetime.datetime.now())[:10]
        DATE_FORMAT = "%Y-%m-%d  %H:%M"
        from_date = datetime.datetime.strptime(datefrm, DATE_FORMAT)
        to_date = datetime.datetime.strptime(dateto, DATE_FORMAT)

        #read subjects of mail with date frame mentioned
        for mailItem in self.getMessages():
            recv = str(mailItem.ReceivedTime)
            new_dt = recv[:19]
            mailRecvTime = datetime.datetime.strptime(new_dt,'%Y-%m-%d %H:%M:%S')

            if mailRecvTime > from_date and mailRecvTime < to_date:
                subj = mailItem.subject.upper()
                lineSplit = subj.split(" ")
                for x in lineSplit[:]:
                    data = re.findall(r'MTS(.*)', x, re.M | re.I)
                    if data:
                        data = 'MTS' + ''.join(data)
                        data1 = pandas.DataFrame({"Case#":[data],"Date": [mailRecvTime]})
                        df = df.append(data1)
	
        #write data to excel using pandas
        if not(df.empty):
            try:
                book = load_workbook(fileName)
                writer = pandas.ExcelWriter(fileName, engine='openpyxl')
                writer.book = book
                df.to_excel(writer,dataExcel+"AlertsQueued",index=False)
                writer.save()
                writer.close()
            except PermissionError:
                print("Could not open file for writing data! Please close this ({}) Excel!".format(fileName))


if __name__ == "__main__":
    processor = EmailReader()
    try:
        inp1 = '2019-06-27 00:01'
        inp2 = '2019-06-27 23:59'
        processor.processEmailRequest(inp1, inp2)
        processor.close()
    except ValueError as value:
        print("The error is ({}). Please provide correct date format".format(value))


